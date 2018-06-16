import googlemaps as ggm
from math import sin, cos, sqrt, atan2, radians
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

api_key = 'AIzaSyDSbtA0_Tz3jt215tYXIOTKArJR5zHWfYI'
gm = ggm.Client(key=api_key)


class Visualizer:
    def draw_node(self, coordinate):
        plt.axis('off')
        plt.plot(coordinate.long, coordinate.lat, 'bo', markersize=12)
        plt.text(coordinate.long, coordinate.lat, coordinate.address, fontsize=6)
        plt.pause(0.05)

    def show(self):
        plt.show()


class Node:
    code = ''
    address = ''
    lat = 0.0
    long = 0.0

    def __init__(self, code, address):

        geocode_result = gm.geocode(address)[0]
        location = geocode_result['geometry']['location']
        self.code = code
        self.address = address
        self.lat = location['lat']
        self.long = location['lng']
        print('get_position(', address, ') = ', self.lat, ' | ', self.long)

    def distance_to(self, another):
        lon1, lat1 = self.long, self.lat
        lon2, lat2 = another.long, another.lat

        R = 6371000  # radius of Earth in meters
        phi_1 = radians(lat1)
        phi_2 = radians(lat2)

        delta_phi = radians(lat2 - lat1)
        delta_lambda = radians(lon2 - lon1)

        a = sin(delta_phi / 2.0) ** 2 + \
            cos(phi_1) * cos(phi_2) * \
            sin(delta_lambda / 2.0) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))

        return R * c


def cal_distance(_nodes):
    node_count = len(nodes)
    matrix = np.zeros((node_count, node_count))
    for i in range(node_count):
        for j in range(node_count):
            matrix[i][j] = matrix[j][i] = _nodes[i].distance_to(_nodes[j])
    return matrix


nodes = []


def load_data(file_name, nodes):
    wb = load_workbook(file_name)
    print(wb.get_sheet_names())
    ws = wb.get_sheet_by_name('address')
    print(ws)

    for row in range(2, 31):
        code = ws.cell(row=row, column=1).value
        address = ws.cell(row=row, column=3).value
        print(code, ' | ', address)

        node = Node(code=code, address=address)

        nodes.append(node)

        visualizer = Visualizer()
        visualizer.draw_node(node)

    visualizer.show()



def __main__():
    load_data('Database.xlsx', nodes)
    graph = cal_distance(nodes)
    print(graph)
    print(nodes)

__main__()